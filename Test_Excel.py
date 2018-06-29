#_*_ coding: utf-8 _*_

import win32com.client
from Test_MariaDB import WrapDB
import math
import pandas as pd
import xlsxwriter
import openpyxl
from xlutils.copy import copy as xl_copy


# Wrap운용팀 DB Connect
db = WrapDB()
db.connet()

# 데이터 Read
data = db.select()
#print(data)

# Wrap운용팀 DB Disconnect
db.disconnect()


if 0:
    import copy
    from openpyxl import load_workbook
    wb = load_workbook(filename='통합지표_류상진_.xlsx', read_only=False, data_only=False)
    ws = wb['데이터_작업']

    all_columns = ws.columns

    dates = None
    values = None
    for idx, column in enumerate(all_columns):
        #print(idx)
        if idx % 2 == 0:
            dates = copy.copy(column)
            #print(type(dates), dir(dates))
            #print (dates[1], len(dates))
        else:
            values = copy.copy(column)
            #print(type(values), dir(values))
            #print(values[1], len(values))

            group = None
            name = None
            ticker = None
            for idx in range(len(dates)):
                if idx == 0:
                    group = values[idx].value
                    print (group)
                elif idx == 1:
                    name = values[idx].value
                    print (name)
                elif idx == 2:
                    ticker = values[idx].value
                    print (ticker)
                elif math.isnan(dates[idx].value) == True:
                    pass
                else:
                    print(str(dates[idx].value)[:10], "\t", values[idx].value, type(values[idx].value))


                if idx > 4:
                    break



if 0:
    # Create a Pandas Excel writer using Openpyxl as the engine.
    writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
    # 주의: 파일이 암호화 걸리면 workbook load시 에러 발생
    writer.book = openpyxl.load_workbook('test.xlsx')
    # Pandas의 DataFrame 클래스를 그대로 이용해서 엑셀 생성 가능
    data.to_excel(writer, sheet_name='sheet2')
    writer.save()

if 0:
    workbook = xlsxwriter.Workbook('test.xlsx')
    worksheet = workbook.add_worksheet()
    workbook.close()


if 0:
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Add()
    ws = wb.Worksheets("Sheet1")
    ws.Cells(1, 1).Value = "hello world"
    wb.SaveAs('test.xlsx')
    excel.Quit()


'''
# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
# Convert the dataframe to an XlsxWriter Excel object.
data.to_excel(writer, sheet_name='Sheet2')
# Close the Pandas Excel writer and output the Excel file.
writer.save()
'''

'''
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open('test.xlsx')
ws = wb.ActiveSheet
print(ws.Cells(1,1).Value)
excel.Quit()
'''

